import openpyxl


def source_excel(source_file):
    # Load the source Excel file
    source_workbook = openpyxl.load_workbook(source_file)
    
    Source_sheet_list = source_workbook.sheetnames
    print("choose the sheet you want to copy from the source file:")
    for i in source_workbook.sheetnames:
        n = Source_sheet_list.index(i) + 1
        print(str(n) + ". " + i)

    source_sheet_user_input = int(input("Enter the number of the sheet you want to copy: "))
    source_selected_sheet = Source_sheet_list[source_sheet_user_input - 1]


    source_sheet = source_workbook[source_selected_sheet]

    source_rows = source_sheet.iter_rows(values_only=True)
    source_headers = next(source_rows) #it gives 1st row

    source_data = list(source_rows)


    print("please give the unique Key:")
    for i in source_headers:
        print(i)

    source_unique_key = input(" enter the Unique key: ")


    source_unique_key_index = source_headers.index(source_unique_key)
    
    return source_unique_key,source_data,source_unique_key_index,source_headers


def target_excel(target_path,source_unique_key,source_data,source_unique_key_index,source_headers):
    t_workbook = openpyxl.load_workbook(target_path)
    target_sheet_list = t_workbook.sheetnames

    print("please choose the target file you want data to be pasted in:")
    for i in target_sheet_list:
        n = target_sheet_list.index(i) + 1
        print(str(n)+ "."+ i)

    target_sheet_user_input =int(input("please enter the sheet no."))

    target_sheet_selected = target_sheet_list[target_sheet_user_input - 1]
    
    target_sheet = t_workbook[target_sheet_selected]

    target_rows = target_sheet.iter_rows(values_only=True)
    target_header = next(target_rows)

    target_data = list(target_rows)

    target_unique_key_index = target_header.index(source_unique_key)

    for i in target_header:
        if  i in source_headers:
            source_header_index = source_headers.index(i)
            target_header_index = target_header.index(i)
            #print(i, "Source index:", source_header_index, "Target index:", target_header_index)


    print("Unique key received in target:",source_unique_key)
    print("Target unique key index:", target_unique_key_index)

    # for i in target_data:

    #     key = i[target_unique_key_index]

    #     for j in source_data:

    #         key2 = j[source_unique_key_index]

    #         if key == key2:

    #             print("MATCH:", key)

    #             for column in target_header:

    #                 if column in source_headers:

    #                     source_index = source_headers.index(column)
    #                     target_index = target_header.index(column)

    #                     print(column,"Source value:", j[source_index],"Target position:", target_index)

    for target_row_number, i in enumerate(target_data, start=2):

        key = i[target_unique_key_index]

        for j in source_data:

            key2 = j[source_unique_key_index]

            if key == key2:

                for column in target_header:

                    if column in source_headers:

                        source_index = source_headers.index(column)
                        target_index = target_header.index(column)

                        target_cell = target_sheet.cell( row=target_row_number,column=target_index + 1)

                        target_cell.value = j[source_index]
                


    
            
#calling functions

source_unique_key,source_data,source_unique_key_index,source_headers = source_excel("/Users/suresh/Desktop/App/ExcelMapper/source.xlsx")

target_excel("/Users/suresh/Desktop/App/ExcelMapper/Target.xlsx",source_unique_key,source_data,source_unique_key_index,source_headers)

    